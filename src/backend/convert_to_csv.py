"""
Simple script to convert Excel file to CSV to avoid openpyxl issues
"""
import pandas as pd
import sys

print("Converting Excel file to CSV...")

try:
    # Try with different engines
    try:
        print("Attempting with default engine...")
        df = pd.read_excel('data/Online_Retail.xlsx', engine=None)
    except:
        try:
            print("Default failed, trying with xlrd engine...")
            df = pd.read_excel('data/Online_Retail.xlsx', engine='xlrd')
        except:
            print("xlrd failed, trying with openpyxl engine with basic options...")
            df = pd.read_excel('data/Online_Retail.xlsx', engine='openpyxl', read_only=True, data_only=True)
    
    print(f"Successfully read {len(df)} rows")
    
    # Save as CSV
    df.to_csv('data/Online_Retail.csv', index=False)
    print("Successfully saved as CSV!")
    
    # Show first few rows
    print("\nFirst 5 rows:")
    print(df.head())
    
    print(f"\nDataset shape: {df.shape}")
    print(f"Columns: {list(df.columns)}")
    
except Exception as e:
    print(f"Error: {e}")
    print("\nTrying alternative method using csv module directly...")
    
    # Alternative: Use openpyxl directly to read and convert
    try:
        from openpyxl import load_workbook
        import csv
        
        wb = load_workbook('data/Online_Retail.xlsx', read_only=True, data_only=True)
        ws = wb.active
        
        with open('data/Online_Retail.csv', 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
        
        print("Successfully converted using openpyxl directly!")
        wb.close()
        
    except Exception as e2:
        print(f"Alternative method also failed: {e2}")
        sys.exit(1)